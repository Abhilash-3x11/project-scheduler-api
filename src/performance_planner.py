from __future__ import annotations

import re
from datetime import date, datetime, timedelta
from io import BytesIO
from pathlib import Path
from typing import Dict, List, Optional, Set, Tuple, Union

from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries

PREFERRED_SHEET_NAMES = ("Performance Planner", "Performance-Plan")
MILESTONE_SHEET_NAMES = ("Project Milestones", "Performance Milestones")


# -----------------------------
# Helpers
# -----------------------------
def normalize_text(value: object) -> str:
    return re.sub(r"\s+", " ", str(value).strip()).lower() if value is not None else ""


def parse_date_value(value: object) -> Optional[date]:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    text = str(value).strip()
    if not text:
        return None

    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass

    try:
        return datetime.fromisoformat(text).date()
    except ValueError:
        return None


def parse_holidays(values: List[object]) -> Set[date]:
    holidays: Set[date] = set()
    for value in values or []:
        if value is None or value == "":
            continue
        if isinstance(value, datetime):
            holidays.add(value.date())
        elif isinstance(value, date):
            holidays.add(value)
        else:
            holidays.add(datetime.strptime(str(value).strip(), "%Y-%m-%d").date())
    return holidays


def is_working_day(d: date, holidays: Set[date]) -> bool:
    return d.weekday() < 5 and d not in holidays


def next_working_day(d: date, holidays: Set[date]) -> date:
    while not is_working_day(d, holidays):
        d += timedelta(days=1)
    return d


def excel_workday(start_date: date, days: int, holidays: Set[date]) -> date:
    """
    Excel-style WORKDAY behavior:
    - days = 0 => return start_date if working day, else next working day
    - days > 0 => move forward by that many working days, excluding start_date
    - days < 0 => move backward by that many working days, excluding start_date
    """
    if days == 0:
        return next_working_day(start_date, holidays)

    current = start_date
    remaining = abs(days)
    step = 1 if days > 0 else -1

    while remaining > 0:
        current += timedelta(days=step)
        if is_working_day(current, holidays):
            remaining -= 1

    return current


def get_sheet_name_from_candidates(wb, candidates: Tuple[str, ...]) -> Optional[str]:
    for name in candidates:
        if name in wb.sheetnames:
            return name
    return None


def get_performance_sheet_name(wb) -> str:
    sheet_name = get_sheet_name_from_candidates(wb, PREFERRED_SHEET_NAMES)
    if sheet_name:
        return sheet_name
    raise ValueError(
        f"Could not find a performance sheet. Expected one of: {', '.join(PREFERRED_SHEET_NAMES)}"
    )


def get_milestone_sheet_name(wb) -> Optional[str]:
    return get_sheet_name_from_candidates(wb, MILESTONE_SHEET_NAMES)


def mark_workbook_for_recalc(wb) -> None:
    """
    Helps Excel recalculate formulas when the workbook is opened.
    """
    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
        wb.calculation.calcMode = "auto"
    except Exception:
        pass


def _clean_input(value: Optional[str]) -> str:
    return value.strip() if value and value.strip() else ""


# -----------------------------
# Formula-based schedule builder
# -----------------------------
def build_schedule_dates(
    project_start_date: date,
    holidays: Set[date],
) -> Tuple[Dict[int, date], Dict[int, date], Dict[int, int]]:
    """
    Recreates the workbook's planned start/end formula chain for rows 2..33.
    Returns:
      planned_start[row], planned_end[row], duration_override[row]
    """
    ps: Dict[int, date] = {}
    pe: Dict[int, date] = {}
    dur_override: Dict[int, int] = {}

    # Row 2
    ps[2] = next_working_day(project_start_date, holidays)

    # Row 3
    ps[3] = ps[2]
    pe[3] = ps[3]

    # Row 4
    ps[4] = excel_workday(pe[3], 1, holidays)
    pe[4] = ps[4]

    # Row 5
    ps[5] = excel_workday(pe[3], 1, holidays)
    pe[5] = excel_workday(ps[5], 15, holidays)

    # Row 6
    ps[6] = excel_workday(pe[3], 5, holidays)
    pe[6] = excel_workday(ps[6], 0, holidays)

    # Row 7
    ps[7] = excel_workday(pe[6], 2, holidays)
    pe[7] = ps[7]

    # Row 8
    ps[8] = excel_workday(pe[6], 3, holidays)
    pe[8] = ps[8]

    # Row 9
    ps[9] = excel_workday(pe[3], 5, holidays)
    pe[9] = ps[9]

    # Row 10
    ps[10] = excel_workday(pe[7], 2, holidays)
    pe[10] = excel_workday(ps[10], 0, holidays)

    # Row 11
    ps[11] = excel_workday(pe[7], 2, holidays)
    pe[11] = excel_workday(ps[11], 5, holidays)

    # Row 12
    ps[12] = pe[7]
    pe[12] = excel_workday(ps[12], 2, holidays)

    # Row 13
    ps[13] = pe[12]
    pe[13] = excel_workday(ps[13], 2, holidays)

    # Row 14
    ps[14] = pe[13]
    pe[14] = ps[14]

    # Row 15
    ps[15] = ps[3] + timedelta(days=8)

    # Row 16
    ps[16] = excel_workday(ps[15], 5, holidays)
    pe[16] = excel_workday(ps[16], 2, holidays)

    # Row 17
    ps[17] = pe[16]
    pe[17] = excel_workday(ps[17], 2, holidays)

    # Row 18
    ps[18] = pe[17]
    pe[18] = excel_workday(ps[18], 2, holidays)

    # Row 19
    ps[19] = excel_workday(pe[14], 3, holidays)
    pe[19] = ps[19]

    # Row 20
    ps[20] = excel_workday(pe[19], 2, holidays)
    pe[20] = excel_workday(ps[20], 0, holidays)

    # Row 21
    ps[21] = excel_workday(pe[20], 3, holidays)
    pe[21] = ps[21]

    # Row 22
    ps[22] = excel_workday(pe[21], 2, holidays)
    pe[22] = excel_workday(ps[22], 0, holidays)

    # Row 23
    ps[23] = excel_workday(pe[22], 3, holidays)
    pe[23] = ps[23]

    # Row 24
    ps[24] = excel_workday(pe[23], 0, holidays)
    pe[24] = ps[24]

    # Row 25
    ps[25] = excel_workday(pe[24], 3, holidays)
    pe[25] = excel_workday(ps[25], 0, holidays)

    # Row 26
    ps[26] = excel_workday(pe[25], 4, holidays)
    pe[26] = excel_workday(ps[26], 0, holidays)

    # Row 27
    ps[27] = excel_workday(pe[26], 2, holidays)
    pe[27] = ps[27]

    # Row 28
    ps[28] = pe[25]
    pe[28] = excel_workday(ps[28], 2, holidays)

    # Row 29
    ps[29] = pe[28]
    pe[29] = pe[27]

    # Row 30
    ps[30] = pe[29]
    pe[30] = excel_workday(ps[30], 2, holidays)

    # Row 31
    ps[31] = pe[27]
    pe[31] = excel_workday(ps[31], 3, holidays)

    # Row 32
    ps[32] = excel_workday(pe[31], 1, holidays)
    pe[32] = excel_workday(ps[32], 2, holidays)

    # Row 33
    ps[33] = excel_workday(pe[32], 0, holidays)
    pe[33] = excel_workday(ps[33], 2, holidays)

    # Summary row overrides
    pe[15] = pe[22] - timedelta(days=1)
    dur_override[15] = (pe[15] - ps[15]).days
    pe[2] = pe[33]

    return ps, pe, dur_override


# -----------------------------
# Owner replacement logic
# -----------------------------
def resolve_owner_from_role_text(
    raw_role_text: object,
    project_manager: str,
    business_consultant: str,
    technical_consultant: str,
    customer: str,
) -> Optional[str]:
    """
    Owner column contains role text.
    Replace with actual owner names supplied by the API inputs.
    """
    if raw_role_text is None or str(raw_role_text).strip() == "":
        return None

    raw = str(raw_role_text).strip()
    key = normalize_text(raw)

    pm = _clean_input(project_manager)
    bc = _clean_input(business_consultant)
    tc = _clean_input(technical_consultant)
    cu = _clean_input(customer)

    if key == "project manager":
        return pm or bc or None

    if key == "business consultant":
        return bc or None

    if key == "technical consultant":
        return tc or None

    if key == "customer":
        return cu or None

    if key in {"customer + business consultant", "business consultant + customer"}:
        parts = []
        if bc:
            parts.append(bc)
        if cu:
            parts.append(cu)
        return "/".join(parts) if parts else None

    tokens = [p.strip() for p in re.split(r"\s*(?:\+|/|,|&| and )\s*", raw) if p.strip()]
    resolved: List[str] = []

    for token in tokens:
        token_key = normalize_text(token)

        if token_key == "project manager":
            value = pm or bc or None
        elif token_key == "business consultant":
            value = bc or None
        elif token_key == "technical consultant":
            value = tc or None
        elif token_key == "customer":
            value = cu or None
        else:
            value = None

        if value and value not in resolved:
            resolved.append(value)

    return "/".join(resolved) if resolved else None


def replace_roles_with_owners(
    ws,
    owner_col: int,
    project_manager: str,
    business_consultant: str,
    technical_consultant: str,
    customer: str,
) -> None:
    for r in range(2, ws.max_row + 1):
        current_value = ws.cell(r, owner_col).value
        resolved = resolve_owner_from_role_text(
            current_value,
            project_manager=project_manager,
            business_consultant=business_consultant,
            technical_consultant=technical_consultant,
            customer=customer,
        )
        if resolved is not None:
            ws.cell(r, owner_col).value = resolved


def remove_technical_consultant_rows_if_needed(
    ws,
    owner_col: int,
    role_col: Optional[int],
    technical_consultant: str,
) -> None:
    """
    If technical consultant input is blank, hide/clear Technical Consultant rows.
    This avoids breaking row-based references in the workbook.
    """
    if _clean_input(technical_consultant):
        return

    rows_to_clear: List[int] = []

    for r in range(2, ws.max_row + 1):
        owner_text = normalize_text(ws.cell(r, owner_col).value)
        role_text = normalize_text(ws.cell(r, role_col).value) if role_col else ""

        if "technical consultant" in owner_text or "technical consultant" in role_text:
            rows_to_clear.append(r)

    for r in rows_to_clear:
        for c in range(1, ws.max_column + 1):
            ws.cell(r, c).value = None
        ws.row_dimensions[r].hidden = True
        ws.row_dimensions[r].height = 0


# -----------------------------
# Milestones formula evaluator
# -----------------------------
_REF_RE = re.compile(
    r"^(?:(?P<sheet>'[^']+'|[^!]+)!)?(?P<cell>\$?[A-Z]{1,3}\$?\d+)$"
)
_RANGE_RE = re.compile(
    r"^(?:(?P<sheet>'[^']+'|[^!]+)!)?(?P<rng>\$?[A-Z]{1,3}\$?\d+:\$?[A-Z]{1,3}\$?\d+)$"
)
_NUM_RE = re.compile(r"^-?\d+(?:\.\d+)?$")


def _unquote_sheet_name(sheet_token: str) -> str:
    sheet_token = sheet_token.strip()
    if sheet_token.startswith("'") and sheet_token.endswith("'"):
        return sheet_token[1:-1]
    return sheet_token


def _strip_dollars(ref: str) -> str:
    return ref.replace("$", "")


def _split_top_level(text: str, separator: str) -> List[str]:
    parts: List[str] = []
    start = 0
    depth = 0
    in_quotes = False
    i = 0

    while i < len(text):
        ch = text[i]
        if ch == '"':
            in_quotes = not in_quotes
        elif not in_quotes:
            if ch == "(":
                depth += 1
            elif ch == ")":
                depth -= 1
            elif depth == 0 and text[i : i + len(separator)] == separator:
                parts.append(text[start:i])
                start = i + len(separator)
                i += len(separator) - 1
        i += 1

    parts.append(text[start:])
    return parts


def _strip_outer_parens(text: str) -> str:
    text = text.strip()
    while text.startswith("(") and text.endswith(")"):
        inner = text[1:-1].strip()
        depth = 0
        in_quotes = False
        balanced = True
        for i, ch in enumerate(inner):
            if ch == '"':
                in_quotes = not in_quotes
            elif not in_quotes:
                if ch == "(":
                    depth += 1
                elif ch == ")":
                    depth -= 1
                    if depth < 0:
                        balanced = False
                        break
        if balanced and depth == 0:
            text = inner
        else:
            break
    return text


def _split_args(arg_text: str) -> List[str]:
    return [part.strip() for part in _split_top_level(arg_text, ",")]


def _parse_number(text: str) -> Optional[Union[int, float]]:
    text = text.strip()
    if not _NUM_RE.match(text):
        return None
    return float(text) if "." in text else int(text)


def _format_date_value(value: object, fmt: str) -> str:
    if isinstance(value, datetime):
        dt = value
    elif isinstance(value, date):
        dt = datetime(value.year, value.month, value.day)
    else:
        parsed = parse_date_value(value)
        if parsed is None:
            return str(value)
        dt = datetime(parsed.year, parsed.month, parsed.day)

    fmt_map = fmt.lower()
    fmt_map = fmt_map.replace("yyyy", "%Y")
    fmt_map = fmt_map.replace("mmmm", "%B")
    fmt_map = fmt_map.replace("mmm", "%b")
    fmt_map = fmt_map.replace("dd", "%d")
    fmt_map = fmt_map.replace("mm", "%m")
    fmt_map = fmt_map.replace("yy", "%y")
    return dt.strftime(fmt_map)


def _resolve_reference_value(
    ref_text: str,
    wb,
    current_sheet_name: str,
    seen: Set[str],
):
    ref_text = ref_text.strip()

    match = _REF_RE.match(ref_text)
    if match:
        sheet_name = _unquote_sheet_name(match.group("sheet")) if match.group("sheet") else current_sheet_name
        cell_ref = _strip_dollars(match.group("cell"))
        ws = wb[sheet_name]
        value = ws[cell_ref].value
        if isinstance(value, str) and value.startswith("="):
            key = f"{sheet_name}!{cell_ref}"
            if key in seen:
                return None
            seen.add(key)
            try:
                return _evaluate_formula(value, wb, sheet_name, seen)
            finally:
                seen.discard(key)
        return value

    match = _RANGE_RE.match(ref_text)
    if match:
        sheet_name = _unquote_sheet_name(match.group("sheet")) if match.group("sheet") else current_sheet_name
        rng = _strip_dollars(match.group("rng"))
        ws = wb[sheet_name]
        min_col, min_row, max_col, max_row = range_boundaries(rng)
        values: List[object] = []
        for row in ws.iter_rows(
            min_row=min_row,
            max_row=max_row,
            min_col=min_col,
            max_col=max_col,
        ):
            for cell in row:
                value = cell.value
                if isinstance(value, str) and value.startswith("="):
                    key = f"{sheet_name}!{cell.coordinate}"
                    if key in seen:
                        continue
                    seen.add(key)
                    try:
                        value = _evaluate_formula(value, wb, sheet_name, seen)
                    finally:
                        seen.discard(key)
                values.append(value)
        return values

    if ref_text.startswith("'") and ref_text.endswith("'"):
        return ref_text[1:-1]

    if ref_text.startswith('"') and ref_text.endswith('"'):
        return ref_text[1:-1]

    num = _parse_number(ref_text)
    if num is not None:
        return num

    if ref_text.upper() == "TRUE":
        return True
    if ref_text.upper() == "FALSE":
        return False

    parsed_date = parse_date_value(ref_text)
    if parsed_date is not None:
        return parsed_date

    return ref_text


def _evaluate_condition(expr: str, wb, current_sheet_name: str, seen: Set[str]) -> bool:
    expr = _strip_outer_parens(expr.strip())

    if expr.upper().startswith("AND(") and expr.endswith(")"):
        inner = expr[4:-1]
        return all(_evaluate_condition(part, wb, current_sheet_name, seen) for part in _split_args(inner))

    if expr.upper().startswith("OR(") and expr.endswith(")"):
        inner = expr[3:-1]
        return any(_evaluate_condition(part, wb, current_sheet_name, seen) for part in _split_args(inner))

    # Comparison operators
    for op in ("<>", ">=", "<=", "=", ">", "<"):
        idx = _find_top_level_operator(expr, op)
        if idx is not None:
            left = expr[:idx].strip()
            right = expr[idx + len(op) :].strip()
            lv = _evaluate_formula(left, wb, current_sheet_name, seen)
            rv = _evaluate_formula(right, wb, current_sheet_name, seen)

            # Normalize dates if possible
            lv_date = parse_date_value(lv) if not isinstance(lv, (date, datetime)) else lv
            rv_date = parse_date_value(rv) if not isinstance(rv, (date, datetime)) else rv

            if isinstance(lv_date, datetime):
                lv_date = lv_date.date()
            if isinstance(rv_date, datetime):
                rv_date = rv_date.date()

            try:
                if op == "=":
                    return lv == rv
                if op == "<>":
                    return lv != rv
                if op == ">":
                    return lv > rv
                if op == "<":
                    return lv < rv
                if op == ">=":
                    return lv >= rv
                if op == "<=":
                    return lv <= rv
            except Exception:
                # Fall back to string comparison
                lhs = "" if lv is None else str(lv)
                rhs = "" if rv is None else str(rv)
                if op == "=":
                    return lhs == rhs
                if op == "<>":
                    return lhs != rhs
                if op == ">":
                    return lhs > rhs
                if op == "<":
                    return lhs < rhs
                if op == ">=":
                    return lhs >= rhs
                if op == "<=":
                    return lhs <= rhs

    # Fallback truthiness
    value = _evaluate_formula(expr, wb, current_sheet_name, seen)
    return bool(value)


def _find_top_level_operator(expr: str, operator: str) -> Optional[int]:
    depth = 0
    in_quotes = False
    i = 0
    while i < len(expr):
        ch = expr[i]
        if ch == '"':
            in_quotes = not in_quotes
        elif not in_quotes:
            if ch == "(":
                depth += 1
            elif ch == ")":
                depth -= 1
            elif depth == 0 and expr[i : i + len(operator)] == operator:
                return i
        i += 1
    return None


def _evaluate_formula(expr: str, wb, current_sheet_name: str, seen: Optional[Set[str]] = None):
    if seen is None:
        seen = set()

    if not isinstance(expr, str):
        return expr

    text = expr.strip()
    if text.startswith("="):
        text = text[1:].strip()

    text = _strip_outer_parens(text)

    if text == "":
        return ""

    # Concatenation
    if "&" in text:
        parts = _split_top_level(text, "&")
        if len(parts) > 1:
            out = []
            for part in parts:
                val = _evaluate_formula(part, wb, current_sheet_name, seen)
                out.append("" if val is None else str(val))
            return "".join(out)

    # Functions
    upper = text.upper()

    if upper.startswith("WORKDAY(") and text.endswith(")"):
        inner = text[len("WORKDAY(") : -1]
        args = _split_args(inner)
        if len(args) < 2:
            return None

        start_val = _evaluate_formula(args[0], wb, current_sheet_name, seen)
        days_val = _evaluate_formula(args[1], wb, current_sheet_name, seen)

        start_date = parse_date_value(start_val)
        if start_date is None and isinstance(start_val, (date, datetime)):
            start_date = start_val.date() if isinstance(start_val, datetime) else start_val

        try:
            days_int = int(float(days_val))
        except Exception:
            return None

        holidays: Set[date] = set()
        if len(args) >= 3:
            hol_val = _evaluate_formula(args[2], wb, current_sheet_name, seen)
            if isinstance(hol_val, list):
                for item in hol_val:
                    if item is None or item == "":
                        continue
                    parsed = parse_date_value(item)
                    if parsed:
                        holidays.add(parsed)
            elif hol_val is not None:
                parsed = parse_date_value(hol_val)
                if parsed:
                    holidays.add(parsed)

        if start_date is None:
            return None

        return excel_workday(start_date, days_int, holidays)

    if upper.startswith("TEXT(") and text.endswith(")"):
        inner = text[len("TEXT(") : -1]
        args = _split_args(inner)
        if len(args) < 2:
            return None
        value = _evaluate_formula(args[0], wb, current_sheet_name, seen)
        fmt = _evaluate_formula(args[1], wb, current_sheet_name, seen)
        if fmt is None:
            return None
        return _format_date_value(value, str(fmt).strip().strip('"'))

    if upper.startswith("DATEVALUE(") and text.endswith(")"):
        inner = text[len("DATEVALUE(") : -1]
        args = _split_args(inner)
        if not args:
            return None
        value = _evaluate_formula(args[0], wb, current_sheet_name, seen)
        return parse_date_value(value)

    if upper.startswith("IF(") and text.endswith(")"):
        inner = text[len("IF(") : -1]
        args = _split_args(inner)
        if len(args) < 2:
            return None
        cond = _evaluate_condition(args[0], wb, current_sheet_name, seen)
        if cond:
            return _evaluate_formula(args[1], wb, current_sheet_name, seen)
        if len(args) >= 3:
            return _evaluate_formula(args[2], wb, current_sheet_name, seen)
        return ""

    if upper.startswith("AND(") and text.endswith(")"):
        inner = text[len("AND(") : -1]
        args = _split_args(inner)
        return all(_evaluate_condition(arg, wb, current_sheet_name, seen) for arg in args)

    if upper.startswith("OR(") and text.endswith(")"):
        inner = text[len("OR(") : -1]
        args = _split_args(inner)
        return any(_evaluate_condition(arg, wb, current_sheet_name, seen) for arg in args)

    # Binary arithmetic
    for op in ("+", "-"):
        idx = _find_top_level_operator(text, op)
        if idx is not None and idx > 0:
            left = text[:idx].strip()
            right = text[idx + 1 :].strip()

            lv = _evaluate_formula(left, wb, current_sheet_name, seen)
            rv = _evaluate_formula(right, wb, current_sheet_name, seen)

            lv_date = parse_date_value(lv) if not isinstance(lv, (date, datetime)) else lv
            rv_date = parse_date_value(rv) if not isinstance(rv, (date, datetime)) else rv

            if isinstance(lv_date, datetime):
                lv_date = lv_date.date()
            if isinstance(rv_date, datetime):
                rv_date = rv_date.date()

            if op == "+":
                if isinstance(lv_date, date) and isinstance(rv, (int, float)):
                    return lv_date + timedelta(days=int(rv))
                if isinstance(lv, (int, float)) and isinstance(rv, (int, float)):
                    return lv + rv
            elif op == "-":
                if isinstance(lv_date, date) and isinstance(rv_date, date):
                    return lv_date - rv_date
                if isinstance(lv_date, date) and isinstance(rv, (int, float)):
                    return lv_date - timedelta(days=int(rv))
                if isinstance(lv, (int, float)) and isinstance(rv, (int, float)):
                    return lv - rv

    # Direct cell/range reference
    if _REF_RE.match(text) or _RANGE_RE.match(text):
        return _resolve_reference_value(text, wb, current_sheet_name, seen)

    # Quoted string literal
    if text.startswith('"') and text.endswith('"'):
        return text[1:-1]

    # Numeric literal
    num = _parse_number(text)
    if num is not None:
        return num

    # Boolean
    if text.upper() == "TRUE":
        return True
    if text.upper() == "FALSE":
        return False

    # Date literal
    parsed_date = parse_date_value(text)
    if parsed_date is not None:
        return parsed_date

    # Fallback: leave as-is
    return text


def populate_milestone_sheet_formulas(wb) -> None:
    """
    Evaluates formulas in the Project Milestones / Performance Milestones sheet.
    The whole sheet is processed so dependencies in other columns are also resolved.
    """
    milestone_sheet_name = get_milestone_sheet_name(wb)
    if not milestone_sheet_name:
        return

    ws = wb[milestone_sheet_name]

    # Multiple passes to resolve dependencies from top to bottom
    for _ in range(8):
        changed = False

        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    result = _evaluate_formula(cell.value, wb, milestone_sheet_name, set())
                    if result is not None and result != cell.value:
                        cell.value = result
                        if isinstance(result, (date, datetime)):
                            cell.number_format = "dd-mm-yyyy"
                        changed = True

        if not changed:
            break


# -----------------------------
# Apply workbook updates
# -----------------------------
def apply_workbook_updates(
    wb,
    project_start_date: str,
    holidays: Optional[List[object]],
    project_manager: str,
    business_consultant: str,
    technical_consultant: str,
    customer: str,
) -> None:
    performance_sheet_name = get_performance_sheet_name(wb)
    ws = wb[performance_sheet_name]

    headers = {normalize_text(c.value): c.column for c in ws[1] if c.value}

    required = {
        "planned start date": "Planned Start Date",
        "planned end date": "Planned End Date",
        "owner": "Owner",
    }
    missing = [label for key, label in required.items() if key not in headers]
    if missing:
        raise ValueError(
            f"Missing required columns in '{ws.title}': {', '.join(missing)}"
        )

    start_col = headers["planned start date"]
    end_col = headers["planned end date"]
    owner_col = headers["owner"]
    role_col = headers.get("role")
    duration_col = headers.get("duration")

    project_start = parse_date_value(project_start_date)
    if project_start is None:
        raise ValueError(
            "Invalid project_start_date. Use YYYY-MM-DD, for example 2026-01-01."
        )

    holiday_set = parse_holidays(holidays or [])
    planned_start, planned_end, duration_overrides = build_schedule_dates(
        project_start, holiday_set
    )

    for row in range(2, ws.max_row + 1):
        if row not in planned_start or row not in planned_end:
            continue

        ws.cell(row=row, column=start_col).value = planned_start[row]
        ws.cell(row=row, column=end_col).value = planned_end[row]

        if duration_col and row in duration_overrides:
            ws.cell(row=row, column=duration_col).value = duration_overrides[row]

    remove_technical_consultant_rows_if_needed(
        ws=ws,
        owner_col=owner_col,
        role_col=role_col,
        technical_consultant=technical_consultant,
    )

    replace_roles_with_owners(
        ws=ws,
        owner_col=owner_col,
        project_manager=project_manager,
        business_consultant=business_consultant,
        technical_consultant=technical_consultant,
        customer=customer,
    )

    populate_milestone_sheet_formulas(wb)
    mark_workbook_for_recalc(wb)


# -----------------------------
# Public functions
# -----------------------------
def update_project_schedule(
    project_start_date: str,
    holidays: Optional[List[object]] = None,
    project_manager: str = "",
    business_consultant: str = "",
    technical_consultant: str = "",
    customer: str = "",
    template_path: Optional[str] = None,
    output_path: Optional[str] = None,
) -> str:
    if template_path is None:
        template_path = str(
            Path(__file__).resolve().parent.parent
            / "data"
            / "Galaxy Elevate Project Workbook - YesPrep.xlsx"
        )

    if output_path is None:
        output_path = str(Path(__file__).resolve().parent.parent / "data" / "output.xlsx")

    wb = load_workbook(template_path)
    apply_workbook_updates(
        wb=wb,
        project_start_date=project_start_date,
        holidays=holidays,
        project_manager=project_manager,
        business_consultant=business_consultant,
        technical_consultant=technical_consultant,
        customer=customer,
    )

    wb.save(output_path)
    return output_path


def update_project_schedule_stream(
    project_start_date: str,
    holidays: Optional[List[object]] = None,
    project_manager: str = "",
    business_consultant: str = "",
    technical_consultant: str = "",
    customer: str = "",
    template_path: Optional[str] = None,
) -> BytesIO:
    if template_path is None:
        template_path = str(
            Path(__file__).resolve().parent.parent
            / "data"
            / "Galaxy Elevate Project Workbook - YesPrep.xlsx"
        )

    wb = load_workbook(template_path)
    apply_workbook_updates(
        wb=wb,
        project_start_date=project_start_date,
        holidays=holidays,
        project_manager=project_manager,
        business_consultant=business_consultant,
        technical_consultant=technical_consultant,
        customer=customer,
    )

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output